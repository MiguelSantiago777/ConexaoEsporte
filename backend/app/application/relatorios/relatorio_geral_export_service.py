"""Exportação do Relatório Geral (todos os polos) pra .xlsx com gráficos
nativos do Excel — mesmos dados e a mesma composição de gráficos que a tela
mostra via Recharts (pizza de beneficiários por polo, barra do ranking de
frequência, linha da evolução semanal), só que editável/interativo no
Excel de verdade, não uma imagem estática."""
import io

import openpyxl
from openpyxl.chart import BarChart, LineChart, PieChart, Reference

from app.application.relatorios.xlsx_estilo import AZUL_MARCA, DOURADO_MARCA, cabecalho_documento, escrever_tabela
from app.interfaces.api.v1.schemas.dashboard_schemas import RelatorioGeralResponse


def exportar_relatorio_geral(relatorio: RelatorioGeralResponse) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório Geral"

    linha = cabecalho_documento(
        ws,
        "Relatório Geral — Todos os Polos",
        f"Período: {relatorio.data_inicio.strftime('%d/%m/%Y')} a {relatorio.data_fim.strftime('%d/%m/%Y')}",
    )

    kpis = escrever_tabela(
        ws,
        ["Polos", "Beneficiários ativos", "Turmas ativas", "Frequência média geral (%)"],
        [[
            relatorio.kpis.total_polos,
            relatorio.kpis.total_beneficiarios_ativos,
            relatorio.kpis.total_turmas_ativas,
            relatorio.kpis.frequencia_media_pct,
        ]],
        linha_inicial=linha,
        titulo="Indicadores",
    )
    linha = kpis.proxima_linha_livre + 1

    ws_polos = wb.create_sheet("Beneficiários por Polo")
    tabela_polos = escrever_tabela(
        ws_polos,
        ["Polo", "Beneficiários"],
        [[s.label, s.valor] for s in relatorio.beneficiarios_por_polo],
        titulo="Beneficiários por Polo",
    )
    if relatorio.beneficiarios_por_polo:
        pizza = PieChart()
        pizza.title = "Beneficiários por Polo"
        dados = Reference(
            ws_polos, min_col=2, min_row=tabela_polos.linha_cabecalho, max_row=tabela_polos.ultima_linha_dados
        )
        categorias = Reference(
            ws_polos, min_col=1, min_row=tabela_polos.primeira_linha_dados, max_row=tabela_polos.ultima_linha_dados
        )
        pizza.add_data(dados, titles_from_data=True)
        pizza.set_categories(categorias)
        pizza.height, pizza.width = 10, 15
        ws_polos.add_chart(pizza, f"E{tabela_polos.linha_cabecalho}")

    ws_ranking = wb.create_sheet("Ranking de Polos")
    tabela_ranking = escrever_tabela(
        ws_ranking,
        ["Polo", "Beneficiários ativos", "Frequência média (%)"],
        [[r.polo_nome, r.beneficiarios_ativos, r.frequencia_media_pct] for r in relatorio.ranking_polos],
        titulo="Ranking de Polos por Frequência",
    )
    if relatorio.ranking_polos:
        barra = BarChart()
        barra.type = "col"
        barra.title = "Frequência média por polo (%)"
        barra.y_axis.scaling.min, barra.y_axis.scaling.max = 0, 100
        dados = Reference(
            ws_ranking, min_col=3, min_row=tabela_ranking.linha_cabecalho, max_row=tabela_ranking.ultima_linha_dados
        )
        categorias = Reference(
            ws_ranking, min_col=1, min_row=tabela_ranking.primeira_linha_dados, max_row=tabela_ranking.ultima_linha_dados
        )
        barra.add_data(dados, titles_from_data=True)
        barra.set_categories(categorias)
        barra.height, barra.width = 10, 18
        serie = barra.series[0]
        serie.graphicalProperties.solidFill = AZUL_MARCA[2:]
        ws_ranking.add_chart(barra, f"F{tabela_ranking.linha_cabecalho}")

    ws_semana = wb.create_sheet("Evolução Semanal")
    tabela_semana = escrever_tabela(
        ws_semana,
        ["Semana", "Frequência (%)"],
        [[s.label, s.valor] for s in relatorio.frequencia_por_semana],
        titulo="Evolução da Frequência Geral por Semana",
    )
    if relatorio.frequencia_por_semana:
        linha_chart = LineChart()
        linha_chart.title = "Evolução da frequência (%)"
        linha_chart.y_axis.scaling.min, linha_chart.y_axis.scaling.max = 0, 100
        dados = Reference(
            ws_semana, min_col=2, min_row=tabela_semana.linha_cabecalho, max_row=tabela_semana.ultima_linha_dados
        )
        categorias = Reference(
            ws_semana, min_col=1, min_row=tabela_semana.primeira_linha_dados, max_row=tabela_semana.ultima_linha_dados
        )
        linha_chart.add_data(dados, titles_from_data=True)
        linha_chart.set_categories(categorias)
        linha_chart.height, linha_chart.width = 10, 18
        serie = linha_chart.series[0]
        serie.graphicalProperties.line.solidFill = DOURADO_MARCA[2:]
        serie.graphicalProperties.line.width = 25000
        ws_semana.add_chart(linha_chart, f"D{tabela_semana.linha_cabecalho}")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
