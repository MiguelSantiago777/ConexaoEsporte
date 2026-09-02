"""
Exportação da Ficha Técnica de Execução da Entidade para .xlsx, no layout
oficial do modelo (Portaria nº 102/2024).

Estratégia: carrega o arquivo-modelo salvo em
`app/infrastructure/templates/ficha_tecnica_execucao.xlsx` (uma cópia
"limpa" do arquivo oficial baixado — layout, mesclas de células e estilos
100% originais, só sem os dados de exemplo que vieram preenchidos nele) e
escreve os valores da Ficha nas células mapeadas abaixo. As coordenadas
foram levantadas inspecionando esse arquivo célula a célula com openpyxl.

Cada polo é sua própria entidade parceira (Termo de Fomento, CNPJ,
representante legal etc. ficam no cadastro do polo) — a seção 1 (dados da
parceria) e a seção 7 (identificação do núcleo) são preenchidas a partir do
polo ao qual a ficha pertence; só a narrativa do período (quantitativo de
beneficiados, modalidades, período de funcionamento, atividades realizadas
e dificuldades) vem da própria ficha.

Limitações conhecidas (documento oficial já é assim, não é bug daqui):
- Reabrir o .xlsx gerado no Excel perde a validação de dropdown
  (Inserido/Não Inserido) da coluna de checklist — é uma limitação do
  openpyxl ao regravar o arquivo; o valor de texto continua correto.
"""
import io
import re
from pathlib import Path

import openpyxl

from app.application.relatorios.cabecalho_convenio import aplicar_cabecalho_xlsx
from app.domain.ficha_execucao.entities import FichaExecucao
from app.domain.polo.entities import Polo

TEMPLATE_PATH = Path(__file__).resolve().parents[2] / "infrastructure" / "templates" / "ficha_tecnica_execucao.xlsx"


def _marcar(marcado: bool) -> str:
    return "(   X   )" if marcado else "(        )"


def _marcar_periodo(texto_original: str, marcado: bool) -> str:
    substituto = "( X )" if marcado else "(    )"
    return re.sub(r"\(\s*\)", substituto, texto_original, count=1)


def _fmt_data(d) -> str:
    return d.strftime("%d/%m/%Y") if d else ""


def exportar_ficha_execucao(
    ficha: FichaExecucao, polo: Polo | None, cabecalho_convenio: str | None = None
) -> io.BytesIO:
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb["Planilha1"]

    # 1 - Dados e informações da parceria (dados do polo/entidade parceira)
    if polo:
        ws["D4"] = polo.processo_sei or ""
        ws["J4"] = polo.termo_fomento_numero or ""
        ws["D5"] = polo.nome_entidade or ""
        ws["D6"] = polo.cnpj or ""
        ws["D7"] = polo.endereco or ""
        ws["D8"] = polo.representante_legal_nome or ""
        ws["K8"] = polo.representante_legal_cpf or ""
        ws["D9"] = polo.objeto or ""
        ws["E10"] = _fmt_data(polo.vigencia_inicio)
        ws["I10"] = _fmt_data(polo.vigencia_fim)
        ws["D11"] = polo.valor_pactuado or ""
        ws["D12"] = polo.valor_executado or ""
        ws["D13"] = polo.parlamentar or ""
        ws["K13"] = polo.emenda or ""

        for aditivo in (polo.termos_aditivos or [])[:2]:
            linha = 16 if aditivo.get("numero", "").upper() == "PRIMEIRO" else 17
            ws[f"C{linha}"] = aditivo.get("objeto", "")
            # data_assinatura vem como string ISO "AAAA-MM-DD" (armazenada em
            # coluna JSON) — reformata para o mesmo padrão dd/mm/aaaa das
            # demais datas do documento.
            data_iso = aditivo.get("data_assinatura") or ""
            partes = data_iso.split("-")
            ws[f"K{linha}"] = f"{partes[2]}/{partes[1]}/{partes[0]}" if len(partes) == 3 else data_iso

    # 1.2 - Ajustes do plano de trabalho
    ws["F19"] = _marcar(ficha.ajuste_status == "NAO_SOLICITADO")
    ws["I19"] = _marcar(ficha.ajuste_status == "APROVADO")
    ws["L19"] = _marcar(ficha.ajuste_status == "NAO_APROVADO")
    ws["A21"] = ficha.ajuste_justificativa or ""

    # 2 - Valores efetivamente recebidos e executados
    valor = ficha.valor_recebido_periodo or ""
    extenso = ficha.valor_recebido_extenso or ""
    ws["A24"] = f"{valor}  {extenso}".strip()
    ws["K24"] = _fmt_data(ficha.data_recebimento)

    # 3 - Análise de valor (2 metas fixas, até 5 etapas cada)
    linhas_por_meta = [range(34, 39), range(41, 46)]
    for meta, linhas in zip(ficha.metas, linhas_por_meta):
        for etapa, linha in zip(meta.get("etapas", []), linhas):
            ws[f"B{linha}"] = etapa.get("nome", "")
            ws[f"G{linha}"] = etapa.get("previsto", "")
            ws[f"I{linha}"] = etapa.get("executado", "")

    # 4 - Desenvolvimento das atividades (comparativo pactuado x executado)
    for item, linha in zip(ficha.atividades_comparativo, range(52, 67)):
        ws[f"E{linha}"] = item.get("pactuado", "")
        ws[f"H{linha}"] = item.get("executado", "")
        ws[f"K{linha}"] = item.get("observacoes", "")

    # 5 - Execução (checklist de documentação)
    for item, linha in zip(ficha.checklist_documentos, range(70, 86)):
        ws[f"G{linha}"] = item.get("situacao", "Não Inserido")
        ws[f"I{linha}"] = item.get("observacao", "")

    # 6 - Inscrição dos beneficiados
    ws["D87"] = _fmt_data(ficha.periodo_inscricao_inicio)
    ws["H87"] = _fmt_data(ficha.periodo_inscricao_fim)
    ws["F88"] = _marcar(ficha.inscricao_todos_nucleos is True)
    ws["H88"] = _marcar(ficha.inscricao_todos_nucleos is False)
    ws["K88"] = ficha.qtd_inscritos if ficha.qtd_inscritos is not None else ""
    ws["D89"] = ficha.observacoes_inscricao or ""

    # 7 - Identificação do núcleo (nome/endereço/responsável vêm do polo;
    # a narrativa do período vem da ficha). O modelo oficial tem um 2º
    # bloco duplicado (linhas 102-111) para quem tem mais de um núcleo por
    # documento — como cada ficha aqui é de um único polo, fica em branco.
    if polo:
        ws["D91"] = polo.nome or ""
        ws["D92"] = polo.endereco or ""
        ws["D93"] = polo.responsavel_nome or ""
        ws["D94"] = polo.responsavel_email or ""
        ws["D95"] = polo.responsavel_telefone or ""
    ws["D96"] = ficha.quantitativo_beneficiados or ""
    ws["D97"] = ficha.modalidades or ""

    periodo = (ficha.periodo_funcionamento or "").upper()
    manha_txt = ws["D98"].value or "(    ) MANHÃ     "
    tarde_txt = ws["G98"].value or "      (    ) TARDE     "
    noite_txt = ws["J98"].value or "  (   ) NOITE     "
    ws["D98"] = _marcar_periodo(manha_txt, "MANHA" in periodo)
    ws["G98"] = _marcar_periodo(tarde_txt, "TARDE" in periodo)
    ws["J98"] = _marcar_periodo(noite_txt, "NOITE" in periodo)

    # A célula de valor é uma única mesclada cobrindo as 2 linhas de rótulo
    # ("Descrição das atividades" / "Dificuldades enfrentadas").
    texto = ficha.descricao_atividades or ""
    if ficha.dificuldades:
        texto = f"{texto}\n\nDificuldades enfrentadas: {ficha.dificuldades}"
    ws["D99"] = texto

    # 9 - Impactos do benefício social
    ws["D113"] = ficha.impactos_sociais or ""
    if ficha.consideracoes_finais:
        ws["D114"] = ficha.consideracoes_finais

    aplicar_cabecalho_xlsx(wb, cabecalho_convenio)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
