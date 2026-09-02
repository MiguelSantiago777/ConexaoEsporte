"""
Exportação da Lista de Presença mensal de uma turma para .xlsx, no layout
oficial do modelo. Carrega o arquivo-modelo em
`app/infrastructure/templates/lista_presenca.xlsx` (cópia fiel do arquivo
oficial baixado) e preenche cabeçalho + grade de presença. As colunas
AJ/AK/AL (Ausências, Presenças, % Presença) já vêm com fórmulas no próprio
modelo — não são tocadas aqui, o Excel recalcula sozinho a partir das
marcações P/A.

Limite do modelo oficial: só cabem 45 beneficiários (linhas 9 a 53) e até
31 dias (colunas D a AH) por planilha.
"""
import calendar
import io
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from app.application.relatorios.cabecalho_convenio import aplicar_cabecalho_xlsx

TEMPLATE_PATH = Path(__file__).resolve().parents[2] / "infrastructure" / "templates" / "lista_presenca.xlsx"

MESES_PT = [
    "", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
]

MAX_BENEFICIARIOS = 45
PRIMEIRA_LINHA_BENEFICIARIO = 9
PRIMEIRA_COLUNA_DIA = 4  # coluna D


def exportar_lista_presenca(
    *,
    nucleo_nome: str,
    turma_descricao: str,
    coordenador_nome: str,
    periodicidade: str,
    professor_nome: str,
    modalidade_nome: str,
    monitor_nome: str,
    horario: str,
    mes: int,
    ano: int,
    beneficiarios: list[tuple[str, str]],  # [(beneficiario_id, nome_completo), ...]
    presencas: dict[tuple[str, int], bool],  # {(beneficiario_id, dia): presente}
    entidade_titulo: str | None = None,  # "LISTA DE PRESENÇA - {entidade} - Termo de Fomento nº {n}"
    cabecalho_convenio: str | None = None,
) -> io.BytesIO:
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb["Lista de Presença"]

    if entidade_titulo:
        ws["B3"] = entidade_titulo
    ws["B4"] = "Núcleo:"  # mantém rótulo original
    ws["D4"] = nucleo_nome
    ws["T4"] = turma_descricao
    ws["D5"] = coordenador_nome
    ws["T5"] = periodicidade
    ws["D6"] = professor_nome
    ws["T6"] = modalidade_nome
    ws["D7"] = monitor_nome
    ws["T7"] = horario
    ws["AB3"] = MESES_PT[mes]
    ws["AB6"] = ano

    dias_no_mes = calendar.monthrange(ano, mes)[1]
    for dia in range(1, 32):
        col = get_column_letter(PRIMEIRA_COLUNA_DIA + dia - 1)
        ws[f"{col}8"] = date(ano, mes, dia) if dia <= dias_no_mes else None
        if dia <= dias_no_mes:
            ws[f"{col}8"].number_format = "DD/MM"

    for idx, (beneficiario_id, nome) in enumerate(beneficiarios[:MAX_BENEFICIARIOS]):
        linha = PRIMEIRA_LINHA_BENEFICIARIO + idx
        ws[f"C{linha}"] = nome
        for dia in range(1, dias_no_mes + 1):
            presente = presencas.get((beneficiario_id, dia))
            if presente is None:
                continue
            col = get_column_letter(PRIMEIRA_COLUNA_DIA + dia - 1)
            ws[f"{col}{linha}"] = "P" if presente else "A"

    aplicar_cabecalho_xlsx(wb, cabecalho_convenio)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
