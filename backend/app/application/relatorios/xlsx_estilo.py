"""Estilo visual compartilhado dos .xlsx gerados pelo sistema — cabeçalho na
cor da marca, bordas finas, largura de coluna automática e painel congelado
no cabeçalho. Usado tanto pelas planilhas tabulares genéricas quanto pelos
relatórios com gráfico nativo (Geral/Polo)."""
from dataclasses import dataclass

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

AZUL_MARCA = "FF00417D"
DOURADO_MARCA = "FFFCBA27"
CINZA_LINHA_ALT = "FFF6F8FA"

FONTE_TITULO = Font(bold=True, size=14, color=AZUL_MARCA)
FONTE_CABECALHO = Font(bold=True, size=11, color="FFFFFFFF")
PREENCHIMENTO_CABECALHO = PatternFill("solid", fgColor=AZUL_MARCA)
PREENCHIMENTO_LINHA_ALT = PatternFill("solid", fgColor=CINZA_LINHA_ALT)
BORDA_FINA = Border(*(Side(style="thin", color="FFD9DEE4") for _ in range(4)))


@dataclass
class TabelaEscrita:
    linha_cabecalho: int
    primeira_linha_dados: int
    ultima_linha_dados: int  # igual a primeira_linha_dados - 1 quando não há nenhuma linha de dado
    proxima_linha_livre: int


def escrever_tabela(
    ws: Worksheet,
    colunas: list[str],
    linhas: list[list],
    *,
    linha_inicial: int = 1,
    titulo: str | None = None,
) -> TabelaEscrita:
    """Escreve uma tabela estilizada a partir da linha indicada (título
    opcional acima do cabeçalho). Retorna as posições da tabela escrita,
    úteis pra montar um gráfico nativo em cima dela."""
    linha = linha_inicial
    if titulo:
        celula = ws.cell(row=linha, column=1, value=titulo)
        celula.font = FONTE_TITULO
        linha += 2

    linha_cabecalho = linha
    for i, nome_coluna in enumerate(colunas, start=1):
        celula = ws.cell(row=linha_cabecalho, column=i, value=nome_coluna)
        celula.font = FONTE_CABECALHO
        celula.fill = PREENCHIMENTO_CABECALHO
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celula.border = BORDA_FINA
    ws.row_dimensions[linha_cabecalho].height = 24

    linha_dados_inicio = linha_cabecalho + 1
    for offset, dados in enumerate(linhas):
        linha_atual = linha_dados_inicio + offset
        for i, valor in enumerate(dados, start=1):
            celula = ws.cell(row=linha_atual, column=i, value=valor)
            celula.border = BORDA_FINA
            if offset % 2 == 1:
                celula.fill = PREENCHIMENTO_LINHA_ALT

    if linhas:
        ws.freeze_panes = ws.cell(row=linha_dados_inicio, column=1)

    _ajustar_largura_colunas(ws, colunas, linhas, coluna_inicial=1)
    return TabelaEscrita(
        linha_cabecalho=linha_cabecalho,
        primeira_linha_dados=linha_dados_inicio,
        ultima_linha_dados=linha_dados_inicio + len(linhas) - 1,
        proxima_linha_livre=linha_dados_inicio + len(linhas) + 1,
    )


def _ajustar_largura_colunas(ws: Worksheet, colunas: list[str], linhas: list[list], coluna_inicial: int) -> None:
    for i, nome_coluna in enumerate(colunas):
        maior = len(str(nome_coluna))
        for dados in linhas:
            if i < len(dados) and dados[i] is not None:
                maior = max(maior, len(str(dados[i])))
        letra = get_column_letter(coluna_inicial + i)
        ws.column_dimensions[letra].width = min(max(maior + 3, 10), 42)


def cabecalho_documento(ws: Worksheet, titulo: str, subtitulo: str | None = None, com_logo: bool = False) -> int:
    """Título grande + subtítulo opcional no topo da planilha. Com
    `com_logo=True`, ancora o logo do projeto em A1 e desloca o texto para
    a coluna C, para não sobrepor a imagem (só usar em planilhas sem
    layout oficial fixo — ver `cabecalho_convenio.py`). Retorna a próxima
    linha livre."""
    coluna = 3 if com_logo else 1
    ws.cell(row=1, column=coluna, value=titulo).font = Font(bold=True, size=16, color=AZUL_MARCA)
    linha = 2
    if subtitulo:
        ws.cell(row=2, column=coluna, value=subtitulo).font = Font(size=10, color="FF6B7280")
        linha = 3
    if com_logo:
        from app.application.relatorios.cabecalho_convenio import logo_flutuante_xlsx

        logo_flutuante_xlsx(ws)
        linha = max(linha, 3)
    return linha + 1


__all__ = [
    "AZUL_MARCA",
    "DOURADO_MARCA",
    "escrever_tabela",
    "cabecalho_documento",
]
