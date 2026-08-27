"""
Exportação da Planilha de Núcleos — RH e Beneficiário para .xlsx, no layout
oficial do modelo. Carrega o arquivo-modelo em
`app/infrastructure/templates/planilha_nucleos.xlsx`.

O modelo oficial duplica o mesmo bloco para "Núcleo 1" e "Núcleo 2" (mesma
lógica "duplicar quantos forem necessários" da Ficha/Grade Horária) — como
aqui a planilha é gerada por polo, só o bloco "Núcleo 1" é preenchido; o
bloco "Núcleo 2" fica em branco.

Vários rótulos do modelo (ex.: "Entidade: ", "Nome do Núcleo/Subnúcleo:")
não têm uma célula de valor separada — o próprio texto do rótulo é
sobrescrito com "rótulo + valor" na mesma célula mesclada, mesmo padrão já
usado no exportador da Ficha Técnica de Execução.

Limites do modelo oficial: até 4 pessoas de RH e até 16 beneficiários no
bloco do Núcleo 1 — excedentes não entram na exportação.
"""
import io
from dataclasses import dataclass
from pathlib import Path

import openpyxl

TEMPLATE_PATH = Path(__file__).resolve().parents[2] / "infrastructure" / "templates" / "planilha_nucleos.xlsx"

PRIMEIRA_LINHA_RH = 9
MAX_RH = 4
PRIMEIRA_LINHA_BENEFICIARIO = 14
MAX_BENEFICIARIOS = 16


@dataclass
class RHItem:
    nome: str
    carga_horaria: str
    telefone: str
    email: str


@dataclass
class BeneficiarioNucleoItem:
    nome: str
    idade: int
    modalidades: str


def exportar_planilha_nucleos(
    *,
    nome_entidade: str,
    termo_fomento_numero: str,
    polo_nome: str,
    polo_horario_funcionamento: str,
    polo_endereco: str,
    rh: list[RHItem],
    beneficiarios: list[BeneficiarioNucleoItem],
) -> io.BytesIO:
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb["Planilha1"]

    ws["A5"] = f"Entidade: {nome_entidade}"
    ws["H5"] = f"Termo de Fomento Nº: {termo_fomento_numero}"
    ws["A7"] = f"Nome do Núcleo/Subnúcleo: {polo_nome}"
    ws["F7"] = f"Dia e horário de funcionamento: {polo_horario_funcionamento}"
    ws["H7"] = f"Endereço do núcleo: {polo_endereco}"

    for idx, pessoa in enumerate(rh[:MAX_RH]):
        linha = PRIMEIRA_LINHA_RH + idx
        ws[f"B{linha}"] = pessoa.nome
        ws[f"F{linha}"] = pessoa.carga_horaria
        ws[f"G{linha}"] = pessoa.telefone
        ws[f"H{linha}"] = pessoa.email

    for idx, beneficiario in enumerate(beneficiarios[:MAX_BENEFICIARIOS]):
        linha = PRIMEIRA_LINHA_BENEFICIARIO + idx
        ws[f"A{linha}"] = beneficiario.nome
        ws[f"E{linha}"] = beneficiario.idade
        ws[f"H{linha}"] = beneficiario.modalidades

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
