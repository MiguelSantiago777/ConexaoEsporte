"""Exportação do Relatório do Polo pra .xlsx com gráficos nativos do Excel
— mesma composição da tela (pizza de beneficiários por modalidade, barra de
frequência por turma, linha da evolução semanal)."""
import io

import openpyxl
from openpyxl.chart import BarChart, LineChart, PieChart, Reference

from app.application.relatorios.xlsx_estilo import AZUL_MARCA, DOURADO_MARCA, cabecalho_documento, escrever_tabela
from app.interfaces.api.v1.schemas.dashboard_schemas import RelatorioPoloResponse


def exportar_relatorio_polo(relatorio: RelatorioPoloResponse) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório do Polo"

    linha = cabecalho_documento(
        ws,
        f"Relatório do Polo — {relatorio.polo_nome}",
        f"Período: {relatorio.data_inicio.strftime('%d/%m/%Y')} a {relatorio.data_fim.strftime('%d/%m/%Y')}",
    )

    kpis = escrever_tabela(
        ws,
        ["Beneficiários ativos", "Turmas ativas", "Frequência média (%)", "Aulas registradas", "Fotos de evidência"],
        [[
            relatorio.kpis.beneficiarios_ativos,
            relatorio.kpis.turmas_ativas,
            relatorio.kpis.frequencia_media_pct,
            relatorio.kpis.aulas_registradas,
            relatorio.kpis.fotos_evidencia,
        ]],
        linha_inicial=linha,
        titulo="Indicadores",
    )
    linha = kpis.proxima_linha_livre

    ws_modalidade = wb.create_sheet("Por Modalidade")
    tabela_modalidade = escrever_tabela(
        ws_modalidade,
        ["Modalidade", "Beneficiários"],
        [[s.label, s.valor] for s in relatorio.beneficiarios_por_modalidade],
        titulo="Beneficiários por Modalidade",
    )
    if relatorio.beneficiarios_por_modalidade:
        pizza = PieChart()
        pizza.title = "Beneficiários por Modalidade"
        dados = Reference(
            ws_modalidade,
            min_col=2,
            min_row=tabela_modalidade.linha_cabecalho,
            max_row=tabela_modalidade.ultima_linha_dados,
        )
        categorias = Reference(
            ws_modalidade,
            min_col=1,
            min_row=tabela_modalidade.primeira_linha_dados,
            max_row=tabela_modalidade.ultima_linha_dados,
        )
        pizza.add_data(dados, titles_from_data=True)
        pizza.set_categories(categorias)
        pizza.height, pizza.width = 10, 15
        ws_modalidade.add_chart(pizza, f"E{tabela_modalidade.linha_cabecalho}")

    ws_turma = wb.create_sheet("Frequência por Turma")
    tabela_turma = escrever_tabela(
        ws_turma,
        ["Turma", "Frequência (%)"],
        [[s.label, s.valor] for s in relatorio.frequencia_por_turma],
        titulo="Frequência por Turma",
    )
    if relatorio.frequencia_por_turma:
        barra = BarChart()
        barra.type = "col"
        barra.title = "Frequência por turma (%)"
        barra.y_axis.scaling.min, barra.y_axis.scaling.max = 0, 100
        dados = Reference(
            ws_turma, min_col=2, min_row=tabela_turma.linha_cabecalho, max_row=tabela_turma.ultima_linha_dados
        )
        categorias = Reference(
            ws_turma, min_col=1, min_row=tabela_turma.primeira_linha_dados, max_row=tabela_turma.ultima_linha_dados
        )
        barra.add_data(dados, titles_from_data=True)
        barra.set_categories(categorias)
        barra.height, barra.width = 10, 18
        serie = barra.series[0]
        serie.graphicalProperties.solidFill = AZUL_MARCA[2:]
        ws_turma.add_chart(barra, f"D{tabela_turma.linha_cabecalho}")

    ws_semana = wb.create_sheet("Evolução Semanal")
    tabela_semana = escrever_tabela(
        ws_semana,
        ["Semana", "Frequência (%)"],
        [[s.label, s.valor] for s in relatorio.frequencia_por_semana],
        titulo="Evolução da Frequência por Semana",
    )
    if relatorio.frequencia_por_semana:
        linha_chart = LineChart()
        linha_chart.title = "Evolução da frequência (%)"
        linha_chart.y_axis.scaling.min, linha_chart.y_axis.scaling.max = 0, 100
        dados = Reference(
            ws_semana, min_col=2, min_row=tabela_semana.linha_cabecalho, max_row=tabela_semana.ultima_linha_dados
        )
        categorias = Reference(
            ws_semana, min_col=1, min_row=tabela_semana.primeira_linha_dados, max_row=tabela_semana.ultima_linha_dados
        )
        linha_chart.add_data(dados, titles_from_data=True)
        linha_chart.set_categories(categorias)
        linha_chart.height, linha_chart.width = 10, 18
        serie = linha_chart.series[0]
        serie.graphicalProperties.line.solidFill = DOURADO_MARCA[2:]
        serie.graphicalProperties.line.width = 25000
        ws_semana.add_chart(linha_chart, f"D{tabela_semana.linha_cabecalho}")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
