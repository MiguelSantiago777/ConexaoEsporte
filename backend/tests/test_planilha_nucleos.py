"""
Testes de edição de Usuário (telefone/carga horária para RH do núcleo) e da
exportação da Planilha de Núcleos — RH e Beneficiário em .xlsx, no layout
oficial do modelo. Cobre:
- RBAC do PATCH /usuarios: MASTER edita qualquer um; GESTOR_POLO só edita
  PROFESSOR do próprio polo, e não pode mudar polo/situação.
- A planilha reflete RH (nome/carga horária/telefone/e-mail) e
  beneficiários (nome/idade/modalidade) do polo.
"""
import io

import openpyxl

from tests.conftest import login


def test_gestor_nao_edita_usuario_de_outro_polo(client, seed_basico):
    token_gestor_a = login(client, "gestor.a@test.com")
    token_gestor_b = login(client, "gestor.b@test.com")
    polo_b_id = str(seed_basico["polo_b"].id)

    professor_b = client.post(
        "/api/v1/usuarios",
        json={
            "nome": "Prof. B", "email": "prof.b@test.com", "senha": "senha123",
            "perfil": "PROFESSOR", "polo_id": polo_b_id,
        },
        headers={"Authorization": f"Bearer {token_gestor_b}"},
    ).json()

    resp = client.patch(
        f"/api/v1/usuarios/{professor_b['id']}",
        json={"telefone": "11999999999"},
        headers={"Authorization": f"Bearer {token_gestor_a}"},
    )
    assert resp.status_code == 403


def test_gestor_nao_pode_mudar_situacao_do_professor(client, seed_basico):
    token_gestor = login(client, "gestor.a@test.com")
    polo_a_id = str(seed_basico["polo_a"].id)
    professor = client.post(
        "/api/v1/usuarios",
        json={
            "nome": "Prof. Situação", "email": "prof.situacao@test.com", "senha": "senha123",
            "perfil": "PROFESSOR", "polo_id": polo_a_id,
        },
        headers={"Authorization": f"Bearer {token_gestor}"},
    ).json()
    resp = client.patch(
        f"/api/v1/usuarios/{professor['id']}",
        json={"ativo": False},
        headers={"Authorization": f"Bearer {token_gestor}"},
    )
    assert resp.status_code == 403


def test_exportar_planilha_nucleos_reflete_rh_e_beneficiarios(client, seed_basico):
    token = login(client, "gestor.a@test.com")
    headers = {"Authorization": f"Bearer {token}"}
    polo_a_id = str(seed_basico["polo_a"].id)
    modalidade_id = str(seed_basico["modalidade"].id)

    token_master = login(client, "master@test.com")
    client.patch(
        f"/api/v1/polos/{polo_a_id}",
        json={"nome_entidade": "Instituto Teste", "termo_fomento_numero": "TF-001"},
        headers={"Authorization": f"Bearer {token_master}"},
    )

    professor = client.post(
        "/api/v1/usuarios",
        json={
            "nome": "Prof. RH", "email": "prof.rh@test.com", "senha": "senha123",
            "perfil": "PROFESSOR", "polo_id": polo_a_id,
        },
        headers=headers,
    ).json()
    resp_patch = client.patch(
        f"/api/v1/usuarios/{professor['id']}",
        json={"telefone": "(11) 98888-7777", "carga_horaria_semanal": "20h"},
        headers=headers,
    )
    assert resp_patch.status_code == 200, resp_patch.text

    turma = client.post(
        "/api/v1/turmas",
        json={
            "polo_id": polo_a_id, "modalidade_id": modalidade_id,
            "horario_inicio": "08:00", "horario_fim": "09:00", "dias_semana": ["SEG"],
            "limite_vagas": 10,
        },
        headers=headers,
    ).json()
    beneficiario = client.post(
        "/api/v1/beneficiarios",
        json={
            "nome_completo": "Beneficiário Núcleo", "data_nascimento": "2000-06-15",
            "documento": "222.333.444-55", "polo_id": polo_a_id,
        },
        headers=headers,
    ).json()
    client.post(
        f"/api/v1/beneficiarios/{beneficiario['id']}/matriculas",
        json={"turma_id": turma["id"]}, headers=headers,
    )

    resp = client.get(f"/api/v1/polos/{polo_a_id}/planilha-nucleos/exportar", headers=headers)
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb["Planilha1"]
    assert ws["A5"].value == "Entidade: Instituto Teste"
    assert ws["H5"].value == "Termo de Fomento Nº: TF-001"
    assert ws["A7"].value == "Nome do Núcleo/Subnúcleo: Polo A"

    # RH inclui o gestor do polo (já existente no seed) + o professor criado
    # neste teste — o gestor entra primeiro (linha 9), o professor na 10.
    assert ws["B9"].value == "Gestor A"
    assert ws["B10"].value == "Prof. RH"
    assert ws["F10"].value == "20h"
    assert ws["G10"].value == "(11) 98888-7777"
    assert ws["H10"].value == "prof.rh@test.com"

    assert ws["A14"].value == "Beneficiário Núcleo"
    assert isinstance(ws["E14"].value, int)
    assert ws["H14"].value == "Judô"
