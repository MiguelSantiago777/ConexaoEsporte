"""
Testes de Polo (dados de parceria/Termo de Fomento) e Ficha de Execução
(Ficha Técnica de Execução da Entidade — Portaria nº 102/2024). Cobre:
- Cada polo é sua própria entidade parceira: RBAC e persistência dos
  campos de Termo de Fomento/CNPJ/representante legal no cadastro do polo.
- RBAC da Ficha de Execução: exclusivo do MASTER, vinculada a um polo.
- Ficha nasce semeada com as listas fixas do modelo.
- Exportação em .xlsx traz os valores cadastrados (do polo e da ficha) nas
  células certas.
"""
import io

import openpyxl

from tests.conftest import login


def test_gestor_nao_edita_dados_de_parceria_do_polo(client, seed_basico):
    token = login(client, "gestor.a@test.com")
    polo_a_id = str(seed_basico["polo_a"].id)
    resp = client.patch(
        f"/api/v1/polos/{polo_a_id}",
        json={"cnpj": "00.000.000/0001-00"},
        headers={"Authorization": f"Bearer {token}"},
    )
    assert resp.status_code == 403  # edição de polo é exclusiva do MASTER


def test_master_cadastra_dados_de_parceria_no_polo(client, seed_basico):
    token = login(client, "master@test.com")
    headers = {"Authorization": f"Bearer {token}"}
    polo_a_id = str(seed_basico["polo_a"].id)

    r1 = client.patch(
        f"/api/v1/polos/{polo_a_id}",
        json={"nome_entidade": "Instituto Teste", "cnpj": "00.000.000/0001-00", "processo_sei": "SEI-123"},
        headers=headers,
    )
    assert r1.status_code == 200, r1.text
    assert r1.json()["nome_entidade"] == "Instituto Teste"
    assert r1.json()["processo_sei"] == "SEI-123"

    r2 = client.patch(
        f"/api/v1/polos/{polo_a_id}",
        json={"termos_aditivos": [{"numero": "PRIMEIRO", "objeto": "Prorrogação"}]},
        headers=headers,
    )
    assert r2.status_code == 200, r2.text
    body = r2.json()
    assert body["nome_entidade"] == "Instituto Teste"  # campo não reenviado permanece
    assert body["termos_aditivos"][0]["numero"] == "PRIMEIRO"


def test_gestor_nao_acessa_fichas_execucao(client, seed_basico):
    token = login(client, "gestor.a@test.com")
    headers = {"Authorization": f"Bearer {token}"}
    polo_a_id = str(seed_basico["polo_a"].id)
    assert client.get("/api/v1/fichas-execucao", headers=headers).status_code == 403
    assert (
        client.post(
            "/api/v1/fichas-execucao",
            json={"polo_id": polo_a_id, "periodo_referencia": "1º Trimestre 2026"},
            headers=headers,
        ).status_code
        == 403
    )


def test_ficha_execucao_nasce_semeada_com_listas_fixas(client, seed_basico):
    token = login(client, "master@test.com")
    headers = {"Authorization": f"Bearer {token}"}
    polo_a_id = str(seed_basico["polo_a"].id)

    r = client.post(
        "/api/v1/fichas-execucao",
        json={"polo_id": polo_a_id, "periodo_referencia": "1º Trimestre 2026"},
        headers=headers,
    )
    assert r.status_code == 201, r.text
    ficha = r.json()

    assert ficha["polo_id"] == polo_a_id
    assert len(ficha["checklist_documentos"]) == 16
    assert all(item["situacao"] == "Não Inserido" for item in ficha["checklist_documentos"])
    assert len(ficha["metas"]) == 2
    assert len(ficha["atividades_comparativo"]) == 15


def test_exportar_ficha_execucao_gera_xlsx_com_dados_do_polo_e_da_ficha(client, seed_basico):
    token = login(client, "master@test.com")
    headers = {"Authorization": f"Bearer {token}"}
    polo_a_id = str(seed_basico["polo_a"].id)

    client.patch(
        f"/api/v1/polos/{polo_a_id}",
        json={
            "nome_entidade": "Instituto Teste", "cnpj": "00.000.000/0001-00", "processo_sei": "SEI-999",
            "responsavel_nome": "Fulana de Tal",
        },
        headers=headers,
    )
    ficha = client.post(
        "/api/v1/fichas-execucao",
        json={"polo_id": polo_a_id, "periodo_referencia": "1º Trimestre 2026"},
        headers=headers,
    ).json()

    patch = client.patch(
        f"/api/v1/fichas-execucao/{ficha['id']}",
        json={
            "valor_recebido_periodo": "R$ 50.000,00",
            "impactos_sociais": "Aumento de 20% na adesão dos beneficiários.",
            "ajuste_status": "APROVADO",
            "descricao_atividades": "Aulas de judô e natação realizadas semanalmente.",
        },
        headers=headers,
    )
    assert patch.status_code == 200, patch.text

    resp = client.get(f"/api/v1/fichas-execucao/{ficha['id']}/exportar", headers=headers)
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb["Planilha1"]
    assert ws["D5"].value == "Instituto Teste"
    assert ws["D4"].value == "SEI-999"
    assert "R$ 50.000,00" in ws["A24"].value
    assert ws["D113"].value == "Aumento de 20% na adesão dos beneficiários."
    assert ws["I19"].value == "(   X   )"  # ajuste APROVADO marcado
    assert ws["F19"].value != "(   X   )"  # NAO_SOLICITADO não marcado
    assert ws["D91"].value == "Polo A"  # identificação do núcleo vem do polo
    assert ws["D93"].value == "Fulana de Tal"
    assert "Aulas de judô e natação" in ws["D99"].value


def test_exportar_lista_presenca_reflete_chamada_lancada(client, seed_basico):
    token = login(client, "gestor.a@test.com")
    headers = {"Authorization": f"Bearer {token}"}
    polo_a_id = str(seed_basico["polo_a"].id)
    modalidade_id = str(seed_basico["modalidade"].id)

    turma = client.post(
        "/api/v1/turmas",
        json={
            "polo_id": polo_a_id, "modalidade_id": modalidade_id,
            "horario_inicio": "08:00", "horario_fim": "09:00", "dias_semana": ["SEG"],
            "limite_vagas": 10, "coordenador_nome": "Coord. Fulana", "monitor_nome": "Monitor Beltrano",
            "periodicidade": "Semanal",
        },
        headers=headers,
    ).json()
    resp_beneficiario = client.post(
        "/api/v1/beneficiarios",
        json={
            "nome_completo": "Beneficiário Presença", "data_nascimento": "2000-01-01",
            "documento": "111.222.333-44", "polo_id": polo_a_id,
        },
        headers=headers,
    )
    assert resp_beneficiario.status_code == 201, resp_beneficiario.text
    beneficiario = resp_beneficiario.json()
    client.post(
        f"/api/v1/beneficiarios/{beneficiario['id']}/matriculas",
        json={"turma_id": turma["id"]}, headers=headers,
    )
    chamada = client.post(
        "/api/v1/frequencias/chamada",
        json={
            "turma_id": turma["id"], "data": "2026-03-10",
            "presencas": [{"beneficiario_id": beneficiario["id"], "presente": True}],
        },
        headers=headers,
    )
    assert chamada.status_code == 201, chamada.text

    resp = client.get(
        f"/api/v1/turmas/{turma['id']}/lista-presenca/exportar", params={"mes": 3, "ano": 2026}, headers=headers
    )
    assert resp.status_code == 200, resp.text

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb["Lista de Presença"]
    assert ws["D4"].value == "Polo A"
    assert ws["D5"].value == "Coord. Fulana"
    assert ws["T5"].value == "Semanal"
    assert ws["D7"].value == "Monitor Beltrano"
    assert ws["AB3"].value == "Março"
    assert ws["AB6"].value == 2026
    assert ws["C9"].value == "Beneficiário Presença"
    assert ws["M9"].value == "P"  # dia 10 do mês -> coluna D (dia 1) + 9 = M


def test_professor_exporta_lista_presenca_da_propria_turma(client, seed_basico):
    token_gestor = login(client, "gestor.a@test.com")
    headers_gestor = {"Authorization": f"Bearer {token_gestor}"}
    polo_a_id = str(seed_basico["polo_a"].id)
    modalidade_id = str(seed_basico["modalidade"].id)

    turma = client.post(
        "/api/v1/turmas",
        json={
            "polo_id": polo_a_id, "modalidade_id": modalidade_id,
            "horario_inicio": "08:00", "horario_fim": "09:00", "dias_semana": ["SEG"],
            "limite_vagas": 10,
        },
        headers=headers_gestor,
    ).json()

    professor = client.post(
        "/api/v1/usuarios",
        json={
            "nome": "Prof. Teste", "email": "prof.teste@test.com", "senha": "senha123",
            "perfil": "PROFESSOR", "polo_id": polo_a_id,
        },
        headers=headers_gestor,
    ).json()
    client.patch(f"/api/v1/turmas/{turma['id']}", json={"professor_id": professor["id"]}, headers=headers_gestor)

    outro_professor = client.post(
        "/api/v1/usuarios",
        json={
            "nome": "Outro Prof.", "email": "outro.prof@test.com", "senha": "senha123",
            "perfil": "PROFESSOR", "polo_id": polo_a_id,
        },
        headers=headers_gestor,
    ).json()

    token_professor = login(client, "prof.teste@test.com")
    resp = client.get(
        f"/api/v1/turmas/{turma['id']}/lista-presenca/exportar",
        params={"mes": 3, "ano": 2026},
        headers={"Authorization": f"Bearer {token_professor}"},
    )
    assert resp.status_code == 200, resp.text

    token_outro = login(client, "outro.prof@test.com")
    resp_negado = client.get(
        f"/api/v1/turmas/{turma['id']}/lista-presenca/exportar",
        params={"mes": 3, "ano": 2026},
        headers={"Authorization": f"Bearer {token_outro}"},
    )
    assert resp_negado.status_code == 403
